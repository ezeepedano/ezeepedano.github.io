
import openpyxl
from decimal import Decimal
from django.core.exceptions import ValidationError
from .models import Product, Category

class InventoryImportService:
    @staticmethod
    def process_import(file, user):
        """
        Reads an Excel file and updates/creates products based on SKU.
        Expected Headers: SKU, Producto, Categoría, Stock Actual, Precio Costo, Precio Venta
        """
        try:
            wb = openpyxl.load_workbook(file)
            ws = wb.active
            
            # Map headers to column indices
            headers = {}
            for cell in ws[1]:
                if cell.value:
                    headers[str(cell.value).strip().lower()] = cell.column - 1

            # Validate required headers
            required_cols = ['sku']
            missing = [col for col in required_cols if col not in headers]
            if missing:
                return {'success': False, 'message': f"Faltan columnas requeridas: {', '.join(missing)}"}

            updated_count = 0
            created_count = 0
            errors = []

            # Iterate rows (skip header)
            for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                sku = row[headers['sku']]
                
                if not sku:
                    continue # Skip empty rows
                
                sku = str(sku).strip().lower()
                
                # Get fields if they exist in columns
                name = row[headers['producto']] if 'producto' in headers and row[headers['producto']] else None
                category_name = row[headers['categoría']] if 'categoría' in headers and row[headers['categoría']] else None
                stock = row[headers['stock actual']] if 'stock actual' in headers and row[headers['stock actual']] is not None else None
                cost = row[headers['precio costo']] if 'precio costo' in headers and row[headers['precio costo']] is not None else None
                price = row[headers['precio venta']] if 'precio venta' in headers and row[headers['precio venta']] is not None else None

                try:
                    product, created = Product.objects.get_or_create(
                        sku=sku,
                        defaults={
                            'user': user,
                            'name': name or "Nuevo Producto",
                            'stock_quantity': stock if stock is not None else 0,
                            'cost_price': cost if cost is not None else 0,
                            'sale_price': price if price is not None else 0,
                            'description': 'Imported via Excel'
                        }
                    )
                    
                    if not created: 
                        # Update existing
                        if name: product.name = name
                        if stock is not None: product.stock_quantity = Decimal(str(stock))
                        if cost is not None: product.cost_price = Decimal(str(cost))
                        if price is not None: product.sale_price = Decimal(str(price))
                        # Ensure user is set (fix for previously orphaned products)
                        if not product.user: product.user = user
                        
                    # Handle Category
                    if category_name:
                        category, _ = Category.objects.get_or_create(
                            name=category_name, 
                            user=user
                        )
                        product.category = category

                    product.save()
                    
                    if created:
                        created_count += 1
                    else:
                        updated_count += 1

                except Exception as e:
                    errors.append(f"Fila {i} (SKU {sku}): {str(e)}")

            return {
                'success': True, 
                'updated': updated_count, 
                'created': created_count, 
                'errors': errors
            }

        except Exception as e:
            return {'success': False, 'message': f"Error procesando archivo: {str(e)}"}

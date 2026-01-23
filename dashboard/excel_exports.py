"""
Exportación de reportes a Excel
"""
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from sales.models import Sale, Customer
from inventory.models import Product
from finance.models import CashMovement, Purchase
from decimal import Decimal
from datetime import datetime


@login_required
def export_sales_to_excel(request):
    """Exportar ventas a Excel con formato profesional"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Ventas"
    
    # Headers con estilo
    headers = ['Fecha', 'Canal', 'Cliente', 'Nro. Orden', 'Monto Total', 'Estado Pago', 'Monto Pagado', 'Saldo']
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    # Obtener ventas
    sales = Sale.objects.filter(user=request.user).select_related('customer').order_by('-date')
    
    # Filtros opcionales
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    channel = request.GET.get('channel')
    
    if date_from:
        sales = sales.filter(date__gte=date_from)
    if date_to:
        sales = sales.filter(date__lte=date_to)
    if channel and channel != 'ALL':
        sales = sales.filter(channel=channel)
    
    # Data rows
    for idx, sale in enumerate(sales, start=2):
        ws.cell(row=idx, column=1, value=sale.date.strftime('%Y-%m-%d %H:%M'))
        ws.cell(row=idx, column=2, value=sale.get_channel_display())
        ws.cell(row=idx, column=3, value=sale.customer.name if sale.customer else 'Sin cliente')
        ws.cell(row=idx, column=4, value=sale.order_id)
        ws.cell(row=idx, column=5, value=float(sale.total))
        ws.cell(row=idx, column=6, value=sale.get_payment_status_display())
        ws.cell(row=idx, column=7, value=float(sale.paid_amount))
        ws.cell(row=idx, column=8, value=float(sale.balance))
    
    # Auto-ajustar columnas
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Totales
    last_row = len(sales) + 2
    ws.cell(row=last_row, column=4, value="TOTALES:").font = Font(bold=True)
    ws.cell(row=last_row, column=5, value=f"=SUM(E2:E{last_row-1})")
    ws.cell(row=last_row, column=7, value=f"=SUM(G2:G{last_row-1})")
    ws.cell(row=last_row, column=8, value=f"=SUM(H2:H{last_row-1})")
    
    # Response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f'ventas_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename={filename}'
    wb.save(response)
    return response


@login_required
def export_customers_to_excel(request):
    """Exportar clientes con estadísticas RFM"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Clientes"
    
    # Headers
    headers = ['Nombre', 'Email', 'Teléfono', 'Total Órdenes', 'Total Gastado', 'Ticket Promedio', 
               'Primera Compra', 'Última Compra', 'Días sin Comprar', 'Segmento RFM']
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    # Data
    customers = Customer.objects.filter(user=request.user).order_by('-stats__total_spent')
    
    for idx, customer in enumerate(customers, start=2):
        stats = customer.safe_stats
        ws.cell(row=idx, column=1, value=customer.name)
        ws.cell(row=idx, column=2, value=customer.email or '')
        ws.cell(row=idx, column=3, value=customer.phone or '')
        ws.cell(row=idx, column=4, value=stats.total_orders if stats else 0)
        ws.cell(row=idx, column=5, value=float(stats.total_spent) if stats else 0)
        ws.cell(row=idx, column=6, value=float(stats.avg_ticket) if stats else 0)
        ws.cell(row=idx, column=7, value=stats.first_order_date.strftime('%Y-%m-%d') if stats and stats.first_order_date else '')
        ws.cell(row=idx, column=8, value=stats.last_order_date.strftime('%Y-%m-%d') if stats and stats.last_order_date else '')
        ws.cell(row=idx, column=9, value=stats.days_since_last_order if stats else 0)
        ws.cell(row=idx, column=10, value=stats.segment if stats else 'Sin Datos')
    
    # Auto-ajustar
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 40)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f'clientes_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename={filename}'
    wb.save(response)
    return response


@login_required
def export_inventory_to_excel(request):
    """Exportar inventario con alertas de stock"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventario"
    
    headers = ['SKU', 'Producto', 'Categoría', 'Stock Actual', 'Precio Costo', 
               'Precio Venta', 'Valor Inventario', 'Estado Stock']
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    products = Product.objects.filter(user=request.user).select_related('category')
    
    for idx, product in enumerate(products, start=2):
        inventory_value = float(product.stock_quantity * product.cost_price)
        
        # Determinar estado
        if product.stock_quantity == 0:
            status = "AGOTADO"
            status_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        elif product.stock_quantity < 10:
            status = "BAJO"
            status_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
        else:
            status = "OK"
            status_fill = None
        
        ws.cell(row=idx, column=1, value=product.sku)
        ws.cell(row=idx, column=2, value=product.name)
        ws.cell(row=idx, column=3, value=product.category.name if product.category else '')
        ws.cell(row=idx, column=4, value=product.stock_quantity)
        ws.cell(row=idx, column=5, value=float(product.cost_price))
        ws.cell(row=idx, column=6, value=float(product.sale_price))
        ws.cell(row=idx, column=7, value=inventory_value)
        status_cell = ws.cell(row=idx, column=8, value=status)
        if status_fill:
            status_cell.fill = status_fill
            status_cell.font = Font(color="FFFFFF", bold=True)
    
    # Auto-ajustar
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 40)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Totales
    last_row = len(products) + 2
    ws.cell(row=last_row, column=6, value="TOTAL INVENTARIO:").font = Font(bold=True)
    ws.cell(row=last_row, column=7, value=f"=SUM(G2:G{last_row-1})")
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f'inventario_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename={filename}'
    wb.save(response)
    return response

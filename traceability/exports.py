"""
Funciones de exportación a Excel para reportes de trazabilidad.
Genera reportes completos de la cadena: Compra → Producción → Venta
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from django.http import HttpResponse
from datetime import datetime
from django.db.models import Sum

from .models import (
    IngredientLot,
    ProductionBatch,
    BatchConsumption,
    SaleBatchAllocation
)


class TraceabilityExporter:
    """Genera reportes de Excel para trazabilidad"""
    
    @staticmethod
    def _apply_header_style(ws):
        """Aplica estilo a la fila de encabezado"""
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
    
    @staticmethod
    def _auto_adjust_columns(ws):
        """Ajusta automáticamente el ancho de las columnas"""
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
    
    @staticmethod
    def full_traceability_report():
        """
        Reporte completo: Todas las compras → producciones → ventas
        """
        wb = Workbook()
        
        # Hoja 1: Lotes de Ingredientes (Compras)
        ws1 = wb.active
        ws1.title = "Compras (Materia Prima)"
        ws1.append([
            'ID Interno', 'Ingrediente', 'Lote Proveedor', 
            'Cantidad Inicial (kg)', 'Cantidad Actual (kg)', 
            'Fecha Recepción', 'Vencimiento', 'Estado'
        ])
        
        for lot in IngredientLot.objects.all().order_by('-received_date'):
            ws1.append([
                lot.internal_id,
                lot.ingredient.name,
                lot.supplier_lot,
               float(lot.quantity_initial),
                float(lot.quantity_current),
                lot.received_date.strftime('%d/%m/%Y'),
                lot.expiration_date.strftime('%d/%m/%Y') if lot.expiration_date else 'N/A',
                'Activo' if lot.is_active else ('Merma' if lot.is_wasted else 'Agotado')
            ])
        
        TraceabilityExporter._apply_header_style(ws1)
        TraceabilityExporter._auto_adjust_columns(ws1)
        
        # Hoja 2: Lotes de Producción
        ws2 = wb.create_sheet("Producciones")
        ws2.append([
            'Lote Producción', 'Producto', 'Cantidad Producida (kg)', 
            'Disponible (kg)', 'Fecha Producción', 'Estado'
        ])
        
        for batch in ProductionBatch.objects.all().order_by('-production_date'):
            ws2.append([
                batch.internal_lot_code,
                batch.product.name,
                float(batch.quantity_produced),
                float(batch.quantity_remaining) if batch.quantity_remaining else 0,
                batch.production_date.strftime('%d/%m/%Y'),
                batch.get_status_display()
            ])
        
        TraceabilityExporter._apply_header_style(ws2)
        TraceabilityExporter._auto_adjust_columns(ws2)
        
        # Hoja 3: Consumos (Producción → Ingredientes)
        ws3 = wb.create_sheet("Consumos por Producción")
        ws3.append([
            'Lote Producción', 'Producto', 'Ingrediente', 
            'Lote MP Usado', 'Lote Proveedor', 'Cantidad Consumida (kg)', 'Merma?'
        ])
        
        for consumption in BatchConsumption.objects.select_related(
            'production_batch__product', 'ingredient', 'ingredient_lot'
        ).all():
            ws3.append([
                consumption.production_batch.internal_lot_code,
                consumption.production_batch.product.name,
                consumption.ingredient.name,
                consumption.ingredient_lot.internal_id,
                consumption.ingredient_lot.supplier_lot,
                float(consumption.quantity_consumed),
                'Sí' if consumption.is_waste else 'No'
            ])
        
        TraceabilityExporter._apply_header_style(ws3)
        TraceabilityExporter._auto_adjust_columns(ws3)
        
        # Hoja 4: Ventas (Producción → Cliente)
        ws4 = wb.create_sheet("Ventas (Lotes Vendidos)")
        ws4.append([
            'Orden Venta', 'Fecha Venta', 'Cliente', 'Producto', 
            'Cantidad (unidades)', 'Lote PT Asignado', 'kg Asignados'
        ])
        
        for allocation in SaleBatchAllocation.objects.select_related(
            'sale_item__sale__customer', 'sale_item__product', 'production_batch'
        ).all():
            ws4.append([
                allocation.sale_item.sale.order_id,
                allocation.sale_item.sale.date.strftime('%d/%m/%Y'),
                allocation.sale_item.sale.customer.name if allocation.sale_item.sale.customer else 'N/A',
                allocation.sale_item.product.name if allocation.sale_item.product else allocation.sale_item.product_title,
                allocation.sale_item.quantity,
                allocation.production_batch.internal_lot_code,
                float(allocation.quantity_allocated)
            ])
        
        TraceabilityExporter._apply_header_style(ws4)
        TraceabilityExporter._auto_adjust_columns(ws4)
        
        return wb
    
    @staticmethod
    def single_batch_report(batch_id):
        """
        Reporte detallado de un lote de producción específico.
        Muestra qué ingredientes se usaron y dónde se vendió.
        """
        batch = ProductionBatch.objects.get(pk=batch_id)
        wb = Workbook()
        ws = wb.active
        ws.title = f"Lote {batch.internal_lot_code}"
        
        # Información del Lote
        ws.append(['INFORMACIÓN DEL LOTE'])
        ws['A1'].font = Font(bold=True, size=14)
        ws.append(['Código Lote:', batch.internal_lot_code])
        ws.append(['Producto:', batch.product.name])
        ws.append(['Cantidad Producida:', f"{batch.quantity_produced} kg"])
        ws.append(['Cantidad Disponible:', f"{batch.quantity_remaining or 0} kg"])
        ws.append(['Fecha Producción:', batch.production_date.strftime('%d/%m/%Y')])
        ws.append(['Estado:', batch.get_status_display()])
        ws.append([''])
        
        # Materias Primas Utilizadas
        ws.append(['MATERIAS PRIMAS UTILIZADAS'])
        ws[f'A{ws.max_row}'].font = Font(bold=True, size=12)
        ws.append(['Ingrediente', 'Lote MP', 'Lote Proveedor', 'Cantidad (kg)', 'Vencimiento'])
        header_row = ws.max_row
        
        for consumption in batch.consumptions.select_related('ingredient', 'ingredient_lot').all():
            ws.append([
                consumption.ingredient.name,
                consumption.ingredient_lot.internal_id,
                consumption.ingredient_lot.supplier_lot,
                float(consumption.quantity_consumed),
                consumption.ingredient_lot.expiration_date.strftime('%d/%m/%Y') if consumption.ingredient_lot.expiration_date else 'N/A'
            ])
        
        # Estilo para encabezado de materias primas
        for cell in ws[header_row]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        
        ws.append([''])
        
        # Ventas
        ws.append(['VENTAS (DÓNDE SE VENDIÓ)'])
        ws[f'A{ws.max_row}'].font = Font(bold=True, size=12)
        ws.append(['Orden', 'Fecha', 'Cliente', 'Producto', 'Cantidad Vendida (kg)'])
        header_row = ws.max_row
        
        for allocation in batch.sales.select_related('sale_item__sale__customer', 'sale_item__product').all():
            ws.append([
                allocation.sale_item.sale.order_id,
                allocation.sale_item.sale.date.strftime('%d/%m/%Y'),
                allocation.sale_item.sale.customer.name if allocation.sale_item.sale.customer else 'N/A',
                allocation.sale_item.product.name if allocation.sale_item.product else allocation.sale_item.product_title,
                float(allocation.quantity_allocated)
            ])
        
        # Estilo para encabezado de ventas
        for cell in ws[header_row]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        
        TraceabilityExporter._auto_adjust_columns(ws)
        
        return wb


# Funciones de vista para descargas
def download_full_traceability_report(request):
    """Descarga reporte completo de trazabilidad"""
    wb = TraceabilityExporter.full_traceability_report()
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f'Trazabilidad_Completa_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename={filename}'
    
    wb.save(response)
    return response


def download_batch_report(request, batch_id):
    """Descarga reporte de un lote específico"""
    wb = TraceabilityExporter.single_batch_report(batch_id)
    batch = ProductionBatch.objects.get(pk=batch_id)
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f'Lote_{batch.internal_lot_code}_{datetime.now().strftime("%Y%m%d")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename={filename}'
    
    wb.save(response)
    return response
